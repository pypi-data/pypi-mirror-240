import logging
from copy import copy

from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.dimensions import RowDimension

logger = logging.getLogger(__name__)


class GeometryBlock:
    def __init__(self, worksheet, cell_range, parent=None):
        # self._sheet = sheet
        self._worksheet = worksheet
        self._range = cell_range
        self._children = list()

        self._parent = parent
        if self._parent:
            self._parent.add_child(self)

    @property
    def cell_range(self):
        return copy(self._range)

    @property
    def min_row(self):
        return self._range.min_row

    @property
    def max_row(self):
        return self._range.max_row

    @property
    def min_col(self):
        return self._range.min_col

    @property
    def max_col(self):
        return self._range.max_col

    @property
    def cols_count(self):
        return self.max_col + 1 - self.min_col

    @property
    def rows_count(self):
        return self.max_row + 1 - self.min_row

    def expand(self, extra_size, displace=False):
        if self._parent:
            self._parent.insert_new_cells(
                CellRange(
                    min_col=self.min_col,
                    min_row=self.max_row + 1,
                    max_col=self.max_col,
                    max_row=self.max_row + 1 + extra_size - 1,
                )
            )

        self._range.expand(
            down=extra_size
        )

        return

        # if not parent -> nothing to move

        assert extra_size >= 0

        if extra_size == 0:
            return

        # process children

        # if displace:

        if self._parent:
            extend_axis_max = self._parent._range.max_col
            extend_axis_min = self._parent._range.min_col

            range_to_move = CellRange(
                min_col=extend_axis_min,
                min_row=self._range.max_row + 1,
                max_col=extend_axis_max,
                max_row=self._parent._range.max_row
            )

            self._parent.expand(extra_size)

        else:
            extend_axis_max = self._worksheet.max_column
            extend_axis_min = self._worksheet.min_column

            range_to_move = CellRange(
                min_col=extend_axis_min,
                min_row=self._range.max_row + 1,
                max_col=extend_axis_max,
                max_row=max(self._worksheet.max_row, self._range.max_row + 1)
            )

        self._worksheet.move_range(
            range_to_move.coord,
            rows=extra_size,
            translate=True,
        )

        logger.debug(
            f"Extending to add {extra_size} extra rows. "
            f"Current range {self._range}. "
            f"Displace range {range_to_move}."
        )
        self._range.expand(
            down=extra_size
        )

    def insert_new_cells(self, cell_range):
        """

        Args:
            cell_range:
        """

        # TODO: displace_direction
        # check coords

        extra_size = cell_range.size["rows"]
        static_axis_min = self.min_col
        static_axis_max = self.max_col

        assert cell_range.min_row > self.min_row

        if cell_range.min_row > self.max_row:
            range_to_move = None
        else:
            range_to_move = CellRange(
                min_col=static_axis_min,
                min_row=cell_range.min_row,
                max_col=static_axis_max,
                max_row=self.max_row
            )

        self.expand(extra_size)

        logger.debug(
            f"inserting block {cell_range}. "
            f"Current range {self._range}. "
            f"Parent {self._parent}. "
            f"Displace range {range_to_move}."
        )

        if range_to_move:
            self.move_cells(range_to_move, row_shift=extra_size)

    def shift(self):
        pass

    def move_cells(self, cell_range, row_shift=0, col_shift=0):
        logger.debug(f"move {cell_range}")
        test_range = copy(cell_range)
        assert self._range.issuperset(test_range), f"{self}, {test_range}"
        test_range.shift(col_shift=col_shift, row_shift=row_shift)
        assert self._range.issuperset(test_range)

        for child in self._children:
            if cell_range.issuperset(child.cell_range):
                # TODO: child.shift() recursive
                child._range.shift(col_shift=col_shift, row_shift=row_shift)
                # print("children", child._children)
                for c in child._children:
                    c._range.shift(col_shift=col_shift, row_shift=row_shift)
            else:
                assert child.cell_range.isdisjoint(cell_range)

        # === material things below

        logger.debug(f"Move {cell_range}, r:{row_shift},c:{col_shift}")
        self._worksheet.move_range(
            cell_range.coord,
            rows=row_shift,
            cols=col_shift,
            translate=True,
        )

        for merged_cell in self._worksheet.merged_cells.ranges:
            if not cell_range.isdisjoint(merged_cell):
                merged_cell.shift(col_shift=col_shift, row_shift=row_shift)

        # TODO: col dimensions

        for i in range(self._worksheet.max_row, cell_range.min_row, -1):
            new_idx = i + row_shift
            self._worksheet.row_dimensions[new_idx] = self._worksheet.row_dimensions[i]
            self._worksheet.row_dimensions[i] = RowDimension(self._worksheet)

    def insert(self, block):
        self.insert_new_cells(block.cell_range)
        self.add_child(block)

    def add_child(self, child):
        assert child not in self._children
        self._children.append(child)
        child._parent = self  # TODO

    def project(self, shift_rows=0, shift_cols=0, parent=None):
        new_range = copy(self._range)
        new_range.shift(row_shift=shift_rows)

        new_block = type(self)(
            self._worksheet,
            new_range,
            parent=parent
        )

        for child in self._children:
            child.project(shift_rows, shift_cols, parent=new_block)

        return new_block

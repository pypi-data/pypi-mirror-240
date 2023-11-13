import copy
from typing import Optional

from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from .cell import CellTemplate, FakeCell


class RangeTemplate:
    def __init__(self, worksheet: Worksheet, cell_range: CellRange, excludes: Optional[list[CellRange]] = None):
        excludes = list(excludes) if excludes is not None else list()

        self._cell_templates = list()
        self._merged_ranges = list()
        self.row_dimensions = dict()
        self.col_dimensions = dict()

        rows = worksheet.iter_rows(
            min_row=cell_range.min_row,
            max_row=cell_range.max_row,
            min_col=cell_range.min_col,
            max_col=cell_range.max_col,
        )

        for row in rows:
            self._cell_templates.append(
                [
                    CellTemplate(c) if not any(c.coordinate in e for e in excludes) else FakeCell(c)
                    for c in row
                ]
            )

        sizes = {len(row) for row in self._cell_templates}
        assert len(sizes) == 1
        self.size_x = sizes.pop()
        self.size_y = len(self._cell_templates)

        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.issubset(cell_range):

                # Сохраняем мерженые диапазоны ячеек в координатах шаблона
                relative_range = CellRange(merged_range.coord)
                relative_range.shift(
                    row_shift=-(cell_range.min_row - 1),
                    col_shift=-(cell_range.min_col - 1)
                )
                self._merged_ranges.append(
                    relative_range
                )

            else:
                assert cell_range.isdisjoint(merged_range), f"joint, but not contains {merged_range.coord} and {cell_range.coord}"

        # Сохраняем информацию о свойствах строк и колонок.
        for row_index in range(cell_range.min_row, cell_range.max_row + 1):
            if row_index in worksheet.row_dimensions:
                relative_index = row_index - cell_range.min_row
                self.row_dimensions[relative_index] = copy.copy(
                    worksheet.row_dimensions[row_index]
                )

        for col_index in range(cell_range.min_col, cell_range.max_col + 1):
            if col_index in worksheet.column_dimensions:
                relative_index = col_index - cell_range.min_col
                self.col_dimensions[relative_index] = copy.copy(
                    worksheet.column_dimensions[col_index]
                )

        # TODO: row dimensions

    def apply(self, worksheet: Worksheet, cell_range: CellRange, context: dict,
              apply_row_styles: bool = False, apply_col_style: bool = False):

        # TODO: check size

        rows = worksheet.iter_rows(
            min_row=cell_range.min_row,
            min_col=cell_range.min_col,
            max_row=cell_range.max_row,
            max_col=cell_range.max_col
        )

        for row, row_template in zip(rows, self._cell_templates):
            for cell, cell_template in zip(row, row_template):
                if cell_template._merged:
                    continue
                cell_template.apply(cell, context)

        for merged_range in self._merged_ranges:
            merged_range = copy.copy(merged_range)
            merged_range.shift(
                col_shift=cell_range.min_col - 1,
                row_shift=cell_range.min_row - 1,
            )
            worksheet.merge_cells(
                merged_range.coord
            )

        if apply_row_styles:
            for relative_index, dimension in self.row_dimensions.items():
                worksheet.row_dimensions[cell_range.min_row + relative_index] = dimension

        if apply_col_style:
            for relative_index, dimension in self.col_dimensions.items():
                worksheet.column_dimensions[cell_range.min_col + relative_index] = dimension

'''
Author       : mati1da
Date         : 2023-05-29 14:38:17
LastEditors  : mati1da
LastEditTime : 2023-06-04 21:47:11
Description  : 
'''


from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import MergedCell
from copy import deepcopy


class excelHelper:

    def __inint__(self):
        pass

    @staticmethod
    def readExcel(filename, sheetname):
        wb = load_workbook(filename)
        sheetObject = wb[sheetname]
        row = sheetObject.max_row
        col = sheetObject.max_column
        data = {"sheetObject": sheetObject,
                "row": row,
                "col": col,
                }
        return data

    @staticmethod
    def writeExcel_listNast(filename, data, title: list):
        wb = Workbook()
        wb.create_sheet('Sheet1')
        ws = wb['Sheet1']
        for i in range(len(title)):
            ws.cell(1, i+1).value = title[i]

        for i in range(len(data)):
            for j in range(len(data[i])):
                ws.cell(i+2, j+1).value = str(data[i][j])

        wb.save(filename)

    @staticmethod
    def merge_cell(data):
        """
            传入一个单元格，如果在合并单元格中，那么就拆分单元格，并且赋值左上角的值。
            注：横向不拆分
        :param: {"sheetObject":sheet对象
                "cell_row":该单元格行数,
                "cell_col":该单元格列数}
        return :{"sheetObject":sheetObject,"YN":True,"pos":[合并单元格的起始行，合并单元格的终止行，合并单元格的起始列，合并单元格的终止列]}   
        return :{"sheetObject":sheetObject,"YN":False,"pos":[]}
        """
        ws = data['sheetObject']
        merge_cell = ws.merged_cells
        cell = ws.cell(data["cell_col"], data["cell_row"])
        if isinstance(cell, MergedCell):
            # TODO: 传入单元格取消合并未完成
            pass
        else:
            return {"sheetObject": ws, "YN": False, "pos": []}

    @staticmethod
    def merge_all_cell(data):
        """
            将所有合并单元格拆分，并赋值
            注：横向合并不会进行拆分赋值
        :param: ws:sheet对象
        return shee对象
        """
        ws = data['sheetObject']

        # 取出所有合并的单元格
        merge_cell = ws.merged_cells
        row = data['row']
        col = data['col']

        cr = []
        # 循环单元格
        for data in merge_cell:
            # 取出行列数
            r1, r2, c1, c2 = data.min_row, data.max_row, data.min_col, data.max_col

            # 横向合并，不会添加到赋值中。
            if r2 - r1 > 0:
                cr.append((r1, r2, c1, c2))

        for r in cr:
            ws.unmerge_cells(
                start_row=r[0], end_row=r[1], start_column=r[2], end_column=r[3])

        for r in cr:
            # 合并单元格左上角的值
            starValue = ws.cell(r[0], r[2]).value

            # 循环合并的单元格包括左上角并赋值
            for endr in range(r[0], r[1]+1):
                for endc in range(r[2], r[3]+1):
                    ws.cell(endr, endc).value = starValue

        return {'sheetObject': ws, "row": row, "col": col}

    @staticmethod
    def writeExcel_sheetObject(data, filePath):
        wb = Workbook()
        wb.create_sheet('Sheet1')
        ws = wb['Sheet1']

        for i in range(data['row']):
            if data['sheetObject'].cell(i+1, 1).value =="":
                break
            for j in range(data['col']):
                ws.cell(i+1, j+1).value = data['sheetObject'].cell(i+1, j+1).value
        wb.save(filePath)




def writeExcel_dd(paper_code_list,filePath):
    wb = Workbook()
    ws = wb.create_sheet("作弊", 0)
    ws = wb.create_sheet("作废", 1)
    ws = wb.create_sheet("缺考", 2)
    del wb['Sheet']
    ws.cell(1, 1).value = "考号"
    for i in range(len(paper_code_list)):
        ws.cell(i + 2, 1).value = paper_code_list[i]
    
    #path_inspect("./data/2023-10-13/生成数据/{}/缺考信息".format(subject))
    wb.save(filePath)
    
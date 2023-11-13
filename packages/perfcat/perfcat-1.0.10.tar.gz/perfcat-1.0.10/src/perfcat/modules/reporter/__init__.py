import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import borders
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.chart.line_chart import LineChart
from openpyxl.chart.reference import Reference

THIN_SIDE = borders.Side(border_style=borders.BORDER_THIN)
BORDER = borders.Border(
    top=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE, bottom=THIN_SIDE
)
TAG_FILL = PatternFill("solid", start_color="c0504d")

TITLE_FILL = PatternFill("solid", start_color="f79646")
CONTENT_FILL = PatternFill("solid", start_color="f2dcdb")


def export(app_name: str, data: dict, filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "报告"
    cur_row = 1

    c_date_label: Cell = ws["A1"]
    c_date_text: Cell = ws["B1"]
    c_pkg_name_label: Cell = ws["A2"]
    c_pkg_name_text: Cell = ws["B2"]

    c_date_label.value = "生成日期"
    c_date_label.fill = TAG_FILL
    c_date_label.border = BORDER

    c_date_text.value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c_date_text.fill = CONTENT_FILL
    c_date_text.border = BORDER

    c_pkg_name_label.value = "包名"
    c_pkg_name_label.fill = TAG_FILL
    c_pkg_name_label.border = BORDER

    c_pkg_name_text.value = app_name
    c_pkg_name_text.fill = CONTENT_FILL
    c_pkg_name_text.border = BORDER

    c_device_info_tag: Cell = ws["A4"]
    c_device_info_tag.fill = TAG_FILL
    c_device_info_tag.value = "设备信息"

    cur_row = 4

    device_info: dict = data["device_info"]

    for index, key in enumerate(device_info.keys()):
        cur_row += 1
        c_title = ws.cell(cur_row, 1, key)
        c_content = ws.cell(cur_row, 2, device_info[key])
        c_title.fill = TITLE_FILL
        c_title.border = BORDER
        c_content.fill = CONTENT_FILL
        c_content.border = BORDER

    cur_row += 1

    # 统计

    # 帧率
    stat_data = {}
    fps_data = data["data"]["FPS"]
    record_length = len(fps_data.keys())
    fps_list = [data["fps"] for data in fps_data.values()]
    # 平均帧率
    fps_avg = sum(fps_list) / record_length
    # 帧率18以上比例
    fps_gt_18 = len(list(filter(lambda fps: fps >= 18, fps_list))) / record_length
    # 帧率25以上比例
    fps_gt_25 = len(list(filter(lambda fps: fps >= 25, fps_list))) / record_length

    # 方差
    fps_var = sum([(fps - fps_avg) ** 2 for fps in fps_list]) / (record_length - 1.0)

    # 降帧频率（每小时 降帧次数）
    fps_drop = (
        len(
            list(
                filter(
                    lambda diff: diff > 8,
                    [
                        abs(fps_list[i] - fps_list[i - 1])
                        for i in range(1, record_length)
                    ],
                )
            )
        )
        / record_length
        * 3600
    )

    # Jank率 （每10分钟卡顿次数）
    fps_jank = (
        sum([data["jank"] for data in fps_data.values()]) / record_length * 10 * 60
    )
    fps_big_jank = (
        sum([data["big_jank"] for data in fps_data.values()]) / record_length * 10 * 60
    )

    stat_data["平均帧率"] = (round(fps_avg), "具体看游戏锁帧的帧率")
    stat_data["帧率18以上比例（%）"] = (round(fps_gt_18 * 100, 2), "90%以上最佳")
    stat_data["帧率25以上比例（%）"] = (round(fps_gt_25 * 100, 2), "80%以上最佳")
    stat_data["帧率方差"] = (round(fps_var, 2), "值越低越稳定，0表示极其稳定")
    stat_data["降帧频率（次/小时）"] = (round(fps_drop), "值越大意味着掉帧越频繁，但是不代表有卡顿")
    stat_data["卡顿率（次/10分钟）"] = (round(fps_jank), "值越大意味着体验越不流畅")
    stat_data["大卡顿（次/10分钟)"] = (round(fps_big_jank), "值越大意味着肉眼可见的打断体验的卡顿明显")

    # FrameTime大于100ms的概率
    fps_frametime_gt_100 = 0
    frametime_count = 0
    for data_value in fps_data.values():
        framtimes: list = data_value["*frametimes"]
        fps_frametime_gt_100 += len(list(filter(lambda ft: ft > 100, framtimes)))
        frametime_count += len(framtimes)

    fps_frametime_gt_100 /= frametime_count

    fps_frametime_gt_100_rate = fps_frametime_gt_100 * 60 * 3600

    stat_data["FrameTime大于100ms的比例（%）"] = (
        round(fps_frametime_gt_100 * 100, 2),
        "帧间隔大于100ms比例越高意味着渲染压力越大\n算法不是很准，仅供参考",
    )
    stat_data["FrameTime大于100ms的频率（次/小时）"] = (
        round(fps_frametime_gt_100_rate, 2),
        "帧间隔大于100ms的概率越高意味着渲染影响到流畅体验\n算法不是很准，仅供参考",
    )

    # 内存
    mem_data = data["data"]["Memory"]
    pss_list = [mem["PSS"] for mem in mem_data.values()]
    swap_list = [mem["SwappedDirty"] for mem in mem_data.values()]
    pss_swap_list = [pss_list[i] + swap_list[i] for i in range(record_length)]

    avg_pss = sum(pss_list) / record_length
    max_pss = max(pss_list)
    max_pss_swap = max(pss_swap_list)

    stat_data["平均内存占用(MB)"] = (round(avg_pss, 2), "平均PSS")
    stat_data["峰值内存占用(MB)"] = (round(max_pss, 2), "峰值PSS")
    stat_data["峰值PSS+SWAP(MB)"] = (round(max_pss_swap, 2), "峰值PSS+SWAP")

    # CPU
    cpu_data = data["data"]["CPU"]
    app_cpu_list = [cpu["AppCPU"] for cpu in cpu_data.values()]
    app_cpu_n_list = [cpu["AppCPUNormalized"] for cpu in cpu_data.values()]
    avg_app_cpu = sum(app_cpu_list) / record_length
    app_cpu_lt_60 = len(list(filter(lambda v: v <= 60, app_cpu_list))) / record_length
    app_cpu_lt_80 = len(list(filter(lambda v: v <= 80, app_cpu_list))) / record_length

    avg_app_cpu_n = sum(app_cpu_n_list) / record_length
    app_cpu_n_lt_60 = (
        len(list(filter(lambda v: v <= 60, app_cpu_n_list))) / record_length
    )
    app_cpu_n_lt_80 = (
        len(list(filter(lambda v: v <= 80, app_cpu_n_list))) / record_length
    )

    stat_data["平均AppCPU占用（%）"] = (round(avg_app_cpu, 2), "")
    stat_data["平均AppCPU小于60%的比例（%）"] = (round(app_cpu_lt_60 * 100, 2), "")
    stat_data["平均AppCPU小于80%的比例（%）"] = (round(app_cpu_lt_80 * 100, 2), "")
    stat_data["平均AppCPU占用（标准化%）"] = (round(avg_app_cpu_n, 2), "")
    stat_data["平均AppCPU小于60%的比例（标准化%）"] = (round(app_cpu_n_lt_60 * 100, 2), "")
    stat_data["平均AppCPU小于80%的比例（标准化%）"] = (round(app_cpu_n_lt_80 * 100, 2), "")

    cur_row += 1
    c_stat_tag: Cell = ws.cell(cur_row, 1)
    c_stat_tag.fill = TAG_FILL
    c_stat_tag.value = "统计"

    for index, key in enumerate(stat_data.keys()):
        cur_row += 1
        c_title = ws.cell(cur_row, 1, key)
        c_title.fill = TITLE_FILL
        c_title.border = BORDER
        c_title.comment = Comment(stat_data[key][1], "perfcat")

        c_content = ws.cell(cur_row, 2, stat_data[key][0])
        c_content.fill = CONTENT_FILL
        c_content.border = BORDER

    # 调整列宽
    adjust_width(ws)

    ws_detail = wb.create_sheet("详细数据")

    records = data["data"]

    record_range = data["record_range"]
    title_row = 1

    table = {}

    for tick in range(record_range[0], record_range[1]):
        table[tick] = {}
        for record_name, record_dict in records.items():
            if not record_dict:
                continue

            record_data = record_dict[tick]
            for title, value in record_data.items():
                if isinstance(value, dict):
                    for k, v in value.items():
                        table[tick][f"{title}{k}"] = v
                elif isinstance(value, list):
                    for index, v in enumerate(value):
                        table[tick][f"{title}{index}"] = v
                else:
                    table[tick][title] = value

    if table:
        titles = list(table.values())[0].keys()
        for col, title in enumerate(titles):
            c_title = ws_detail.cell(title_row, col + 2, title)
            c_title.fill = TITLE_FILL
            c_title.border = BORDER

        c_title_sec = ws_detail.cell(title_row, 1, "秒")
        c_title_sec.fill = TITLE_FILL
        c_title_sec.border = BORDER

        for row, row_data in enumerate(table.values()):
            cell = ws_detail.cell(title_row + row + 1, 1, row)  # 写入秒
            cell.fill = CONTENT_FILL
            cell.border = BORDER
            for col, col_name in enumerate(row_data.keys()):
                cell = ws_detail.cell(title_row + row + 1, col + 2, row_data[col_name])
                cell.fill = CONTENT_FILL
                cell.border = BORDER

        adjust_width(ws_detail)

        # FPS chart
        selected_titles = ["fps", "jank", "big_jank"]
        chart = create_chart(ws_detail, selected_titles, titles, "FPS", "")
        ws.add_chart(chart, "A44")

        # CPU chart
        selected_titles = ["AppCPU", "TotalCPU"]
        chart = create_chart(ws_detail, selected_titles, titles, "CPU占用", "%")
        ws.add_chart(chart, "A74")

        # MEM chart
        selected_titles = [
            "PSS",
            "PrivateDirty",
            "PrivateClean",
            "SwappedDirty",
            "HeapSize",
            "HeapAlloc",
            "HeapFree",
        ]
        chart = create_chart(ws_detail, selected_titles, titles, "内存占用", "%")
        ws.add_chart(chart, "A104")

        # Temp chart
        selected_titles = ["整体温度", "CPU温度", "GPU温度", "NPU温度", "电池温度"]
        chart = create_chart(ws_detail, selected_titles, titles, "温度", "%")
        ws.add_chart(chart, "A134")

    wb.save(filename)
    wb.close()


def adjust_width(ws: Worksheet):
    for col in ws.columns:
        col_name = col[0].column_letter
        max_length = max([len(str(cell.value)) for cell in col])
        adjust_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_name].width = adjust_width


def create_chart(
    ws_detail: Worksheet,
    selected_titles: List[str],
    titles: List[str],
    chart_name: str,
    y_axis_name: str,
):
    chart = LineChart()
    chart.title = chart_name
    chart.legend.position = "r"
    chart.y_axis.title = y_axis_name
    chart.width = 40
    chart.height = 12
    titles = list(titles)
    cols = [titles.index(selected) + 2 for selected in selected_titles]

    for col in cols:
        data_ref = Reference(
            ws_detail, min_col=col, max_col=col, min_row=1, max_row=ws_detail.max_row
        )
        chart.add_data(data_ref, titles_from_data=True)

    return chart
